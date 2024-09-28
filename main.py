import numpy as np
from loguru import logger
import os
import easygui
from openpyxl import load_workbook

# 获取脚本所在目录
script_dir = os.path.dirname(os.path.abspath(__file__))
logger.info(f"脚本所在目录: {script_dir}")

# 构建Excel文件的绝对路径
file_path = os.path.join(script_dir, "excel", "永劫无间-征神之路-灵玦图鉴-收集情况.xlsx")
logger.info(f"尝试读取的文件路径: {file_path}")

# 检查文件是否存在
if not os.path.exists(file_path):
    logger.error(f"文件不存在: {file_path}")
    exit(1)

try:
    # 步骤1: 读取excel文件
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]

    # 步骤2: 将工作表数据转换为列表
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]

    # 显示原始数据的前几行
    logger.info("原始数据:")
    for column in columns:
        logger.info(column)
    logger.info(rows[:5])

    while True:
        name = easygui.enterbox(msg="请输入灵玦名称，例如：引雷·青", title="输入灵玦名称")

        if name:
            matched_cells = []
            for row_idx, row in enumerate(rows, start=2):  # 从第2行开始，因为第1行是标题
                for col_idx, cell in enumerate(row, start=1):
                    if isinstance(cell, str) and name in cell:
                        matched_cells.append((row_idx, col_idx))

            if matched_cells:
                uncollected = []
                all_collected = True

                for row, col in matched_cells:
                    exist_value = sheet.cell(row=row, column=col+1).value
                    if exist_value != "✓":
                        uncollected.append(sheet.cell(row=row, column=col).value)
                        all_collected = False

                if all_collected:
                    easygui.msgbox(msg="该灵玦已全部收集，请选择其他灵玦！", title="提示")
                else:
                    uncollected_str = "，".join(uncollected)
                    if easygui.ynbox(msg=f'"{uncollected_str}"未收集，可以选择当前灵玦！\n是否收集？', title="提示"):
                        # 根据未收集的灵玦数量选择不同的选择方式
                        if len(uncollected) == 1:
                            selected_uncollected = easygui.ynbox(msg=f"是否收集 {uncollected[0]}？", title="选择灵玦")
                            selected_uncollected = [uncollected[0]] if selected_uncollected else []
                        else:
                            selected_uncollected = easygui.multchoicebox(msg="请选择要收集的灵玦：", title="选择灵玦", choices=uncollected)

                        if selected_uncollected:
                            for row, col in matched_cells:
                                if sheet.cell(row=row, column=col).value in selected_uncollected:
                                    sheet.cell(row=row, column=col+1, value="✓")
                                    logger.info(f"{sheet.cell(row=row, column=col).value} 已标记为收集")

                            # 保存更改到Excel文件中
                            workbook.save(file_path)
                            logger.info("\n处理后的数据已保存到 '永劫无间-征神之路-灵玦图鉴-收集情况.xlsx'")
                        else:
                            logger.info("用户未选择任何灵玦，程序继续")
            else:
                easygui.msgbox(msg=f"未找到 {name}，请确认输入是否正确！", title="提示")
        else:
            logger.info("用户取消输入，程序退出")
            break

except Exception as e:
    logger.error(f"处理Excel文件时发生错误: {str(e)}")