import os
from loguru import logger
import PySimpleGUI as sg
from openpyxl import load_workbook

# 获取脚本所在目录
script_dir = os.path.dirname(os.path.abspath(__file__))
logger.info(f"脚本所在目录: {script_dir}")

# 构建Excel文件的绝对路径
file_path = os.path.join(script_dir, "excel", "永劫无间-征神之路-灵玦图鉴-收集情况.xlsx")
logger.info(f"尝试读取的文件路径: {file_path}")

# 检查文件是否存在
if not os.path.exists(file_path):
    logger.error(f"文件不存在: {file_path}")
    exit(1)

try:
    # 步骤1: 读取excel文件
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]

    # 步骤2: 获取A-T列的数据
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=20):
        row_data = [cell.value for cell in row]
        if any(row_data):  # 确保行不是空的
            data.append(row_data)

    # 步骤3: 找出未收集的灵玦，并按列组织
    uncollected = {}
    for col in range(0, len(data[0]), 2):
        column_type = data[0][col]
        column_data = []
        for row in data[1:]:
            if row[col] and row[col+1] != "✓":
                column_data.append((row[col], col))
        if column_data:
            uncollected[column_type] = column_data

    # 步骤4: 创建GUI
    layout = [[sg.Text("未收集的灵玦：")]]
    for column_type, column_data in uncollected.items():
        layout.append([sg.Text(f"--- {column_type} ---", font=('Helvetica', 10, 'bold'))])
        layout.append([sg.Checkbox(item[0], key=f'-CHECK-{item[1]}-{i}') for i, item in enumerate(column_data)])
    layout.append([sg.Button('确定'), sg.Button('取消')])

    window = sg.Window('灵玦收集', layout, resizable=True)

    # 步骤5: 事件循环
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == '取消':
            break
        elif event == '确定':
            selected = []
            for key, value in values.items():
                if key.startswith('-CHECK-') and value:
                    col, i = map(int, key.split('-')[2:])
                    selected.append((uncollected[data[0][col]][int(i)][0], col))
            
            if selected:
                for name, col in selected:
                    for row in range(2, sheet.max_row + 1):
                        if sheet.cell(row=row, column=col+1).value == name:
                            sheet.cell(row=row, column=col+2, value="✓")
                            logger.info(f"{name} 已标记为收集")
                workbook.save(file_path)
                sg.popup("选中的灵玦已标记为已收集！")
                break
            else:
                sg.popup("请至少选择一个灵玦！")

    window.close()

except Exception as e:
    logger.error(f"处理Excel文件时发生错误: {str(e)}")